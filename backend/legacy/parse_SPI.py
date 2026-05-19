import re
from datetime import datetime
import re
import copy
import openpyxl
import shutil
from pathlib import Path

# 预编译正则表达式以提高性能
row_pattern = re.compile(r'(\$?[A-Za-z]+\$?)(\d+)')
# 默认配置的根目录（SIS_Agent/parse_spi）
DEFAULT_PARSE_SPI_ROOT = Path(__file__).parent / "parse_spi"
# 从配置文件读取支持的报文类型
def load_message_types(config_path):
    """
    从配置文件读取支持的报文类型和帧头信息
    :param config_path: spi_id.txt 文件的路径
    :return: (message_types, header_message_map)
    """
    message_types = {}
    header_message_map = {}
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                parts = line.split(':')
                if len(parts) > 1:
                    header = parts[0].strip()
                    message_ids = [msg_id.strip() for msg_id in parts[1].split(',') if msg_id.strip()]
                    if header and message_ids:
                        header_message_map[header] = message_ids
                        for msg_id in message_ids:
                            message_types[msg_id] = f'{msg_id}报文'
        if message_types:
            return message_types, header_message_map
        print(f"配置文件格式错误，未找到有效报文类型")
    except Exception as e:
        print(f"读取配置文件时出错: {e}")
    return {}, {}

def ensure_user_has_default_files(workspace_dir, default_root=DEFAULT_PARSE_SPI_ROOT):
    """
    确保用户的 workspace 中有 parse_spi 所需的配置和模板文件
    如果没有，则从默认位置复制
    """
    workspace_dir = Path(workspace_dir)

    # 配置文件 spi_id.txt
    default_config = default_root / "settings" / "spi_id.txt"
    # user_config = workspace_dir / "spi_id.txt"
    # if not user_config.exists() and default_config.exists():
    #     shutil.copy2(default_config, user_config)
    #     print(f"已复制默认配置文件到 {user_config}")

    # 模板文件 template.xlsx
    default_template = default_root / "template" / "template.xlsx"
    # user_template = workspace_dir / "template.xlsx"
    # if not user_template.exists() and default_template.exists():
    #     shutil.copy2(default_template, user_template)
    #     print(f"已复制默认模板文件到 {user_template}")

    return default_config, default_template

def extract_messages(log_file_path, message_type, header=None):
    """
    从SPI日志文件中提取指定类型的报文内容
    :param log_file_path: 日志文件路径
    :param message_type: 报文类型（如'61'、'bb'等）
    :param header: 可选，帧头（如'55ee'、'ee55'等），用于过滤特定帧头的报文
    :return: 提取后的去重报文列表（包含时间和完整报文）
    """
    # 定义正则表达式模式
    # 1. 匹配指定类型报文的标识行（包含funcid :message_type），支持send和recv两种类型
    # 进一步简化正则表达式，使其更加灵活
    funcid_pattern = re.compile(
        fr'^(\d{{4}}-\d{{2}}-\d{{2}} \d{{2}}:\d{{2}}:\d{{2}}\.\d{{3}}),pid:\d+-tid:\d+,D,msg:.*spi service (send|recv) data funcid :{message_type}$'
    )
    # 2. 匹配后续的报文数据行（I级别日志，包含16进制数据）
    data_line_pattern = re.compile(
        r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3}),pid:\d+-tid:\d+,I,msg:([0-9a-fA-F\s]+)$'
    )
    
    extracted_data = []  # 使用列表存储所有找到的报文，不再使用OrderedDict去重
    current_time = None
    current_message = []
    current_funcid = None
    
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()  # 去除行首尾空白字符
                if not line:
                    continue  # 跳过空行
                
                # 检查是否是指定类型报文的标识行
                funcid_match = funcid_pattern.match(line)
                if funcid_match:
                    
                    # 如果已经有正在收集的报文，先保存它
                    if current_time and current_message:
                        # 检查是否符合帧头要求
                        if header:
                            full_message = ''.join(current_message)
                            if full_message.startswith(header):
                                extracted_data.append((current_time, full_message, message_direction))
                        else:
                            full_message = ''.join(current_message)
                            extracted_data.append((current_time, full_message, message_direction))
                    
                    # 记录当前报文的时间戳、funcid和消息类型（send或recv）
                    current_time = funcid_match.group(1)
                    current_funcid = message_type
                    message_direction = funcid_match.group(2)  # 获取是send还是recv
                    current_message = []  # 重置当前报文数据列表
                    continue
                
                # 如果已经捕获到报文标识，开始收集后续数据行
                if current_time and current_message is not None:
                    data_match = data_line_pattern.match(line)
                    if data_match:
                        # 提取16进制数据部分，去除空格
                        hex_data = data_match.group(2).replace(' ', '').strip()
                        if hex_data:
                            # 如果当前报文还没有数据，且指定了帧头，检查是否以该帧头开头
                            if header and not current_message:
                                if not hex_data.startswith(header):
                                    # 如果是第一行数据且不符合帧头要求，重置状态并跳过
                                    current_time = None
                                    current_message = None
                                    current_funcid = None
                                    message_direction = None
                                    continue
                            # 否则（不是第一行或符合帧头要求）添加数据
                            current_message.append(hex_data)
                    else:
                        # 如果遇到非数据行，说明当前报文收集结束
                        if current_message:
                            # 拼接完整报文
                            full_message = ''.join(current_message)
                            
                            # 检查是否符合帧头要求
                            if header:
                                if full_message.startswith(header):
                                    extracted_data.append((current_time, full_message, message_direction))
                            else:
                                extracted_data.append((current_time, full_message, message_direction))
                        # 重置状态
                        current_time = None
                        current_message = []
                        current_funcid = None
                        message_direction = None
        
        # 处理文件末尾可能残留的报文
        if current_time and current_message:
            # 拼接完整报文
            full_message = ''.join(current_message)
            
            # 检查是否符合帧头要求
            if header:
                if full_message.startswith(header):
                    extracted_data.append((current_time, full_message, message_direction))
            else:
                extracted_data.append((current_time, full_message, message_direction))
    except Exception as e:
        print(f"处理文件 {log_file_path} 时出错：{e}")
    
    # 构建报文内容列名
    content_column = f'{message_type}报文内容'
    
    # 转换为列表格式（时间+完整报文）
    result_list = []
    # 使用字典去重
    unique_messages = {}
    for time, full_msg, direction in extracted_data:
        # 创建复合键（时间戳+报文内容）来确保唯一性
        unique_key = f"{time}_{full_msg}"
        if unique_key not in unique_messages:
            # 保持原始格式，将方向信息单独存储
            unique_messages[unique_key] = {
                '时间': time,
                content_column: f"{time},{full_msg}",
                '方向': direction  # 单独存储方向信息
            }
    
    # 转换为列表并按时间排序
    result_list = list(unique_messages.values())
    result_list.sort(key=lambda x: datetime.strptime(x['时间'], '%Y-%m-%d %H:%M:%S.%f'))
    
    # 按时间升序排序（基于datetime对象比较）
    result_list.sort(key=lambda x: datetime.strptime(x['时间'], '%Y-%m-%d %H:%M:%S.%f'))
    
    return result_list

def save_to_excel(messages_dict, excel_path, supported_message_types, template_path=None):
    """
    将提取结果保存到Excel文件
    :param messages_dict: 字典，键为报文类型，值为提取后的报文列表
    :param excel_path: 输出Excel路径
    :param template_path: Excel模板路径
    """
    global openpyxl  # 确保使用全局导入的openpyxl模块
    
    # 检查是否有数据
    if not messages_dict or all(not messages for messages in messages_dict.values()):
        print("没有提取到报文数据")
        return
    
    # 处理所有报文类型的数据
    processed_messages = {}
    # 初始化一个字典，用于按msg_type和direction存储processed_item
    messages_by_direction = {}
    for msg_type in supported_message_types:
        processed_messages[msg_type] = []
        # 为每种报文类型初始化三个方向的列表
        messages_by_direction[msg_type] = {
            'recv': [],
            'send': [],
            'unknown': []
        }
    all_messages = []
    
    # 遍历所有报文类型进行处理
    for msg_type, messages in messages_dict.items():
        # 获取该报文类型的名称
        msg_name = supported_message_types.get(msg_type, f'{msg_type}报文')
        content_key = f'{msg_type}报文内容'
        
        # 遍历该类型的所有报文
        for item in messages:
            # 拆分时间戳和报文内容
            time_stamp = item['时间']
            message_content = item[content_key].split(',', 1)[1] if ',' in item[content_key] else item[content_key]
            
            # 判断报文方向
            # 初始化方向为unknown
            final_direction = 'unknown'
            # 根据前缀判断方向
            if message_content.startswith('55ee'):
                final_direction = 'recv'
            elif message_content.startswith('ee55'):
                final_direction = 'send'
            
            # 构建带方向后缀的sheet_name
            if final_direction == 'recv':
                sheet_name = f'{msg_name}_recv'
            elif final_direction == 'send':
                sheet_name = f'{msg_name}_send'
            else:
                sheet_name = msg_name
            
            # 创建processed_item
            processed_item = {
                '时间戳': time_stamp,
                content_key: message_content,
                'message_type': msg_type,
                'sheet_name': sheet_name,
                '报文种类': msg_name,
                '方向': final_direction
            }
            
            # 添加到processed_messages
            if msg_type in processed_messages:
                processed_messages[msg_type].append(processed_item)
            
            # 添加到messages_by_direction
            if msg_type in messages_by_direction and final_direction in messages_by_direction[msg_type]:
                messages_by_direction[msg_type][final_direction].append(processed_item)
            
            # 添加到all_messages列表用于创建目录
            all_messages.append(processed_item)
    
    if template_path:
        try:
            # 读取模板文件
            print(f"正在使用模板文件：{template_path}")
            # openpyxl 已在文件顶部导入
            # 加载模板文件
            workbook = openpyxl.load_workbook(template_path)
            
            # 存储所有报文工作表
            message_sheets = {}
            
            # 检查是否存在模板sheet
            template_sheet_name = None
            for name in workbook.sheetnames:
                if '{$0}报文' in name:
                    template_sheet_name = name
                    break
            
            # 创建message_sheets字典，用于存储每种报文类型和方向对应的工作表对象
            message_sheets = {}
            
            # 遍历所有报文类型和方向，只创建有数据的工作表
            for msg_type in supported_message_types:
                msg_name = supported_message_types[msg_type]
                
                # 为每种报文类型的每个方向创建工作表，但只在该方向有数据时才创建
                for direction in ['recv', 'send', 'unknown']:
                    # 检查该方向是否有数据
                    if msg_type in messages_by_direction and direction in messages_by_direction[msg_type] and messages_by_direction[msg_type][direction]:
                        # 构建带方向后缀的工作表名称
                        if direction == 'unknown':
                            sheet_title = msg_name
                        else:
                            sheet_title = f'{msg_name}_{direction}'
                        
                        # 查找是否已经存在该工作表
                        found_sheet = None
                        for name in workbook.sheetnames:
                            if name == sheet_title:
                                found_sheet = workbook[name]
                                break
                        
                        # 如果不存在，复制模板工作表
                        if not found_sheet:
                            if template_sheet_name:
                                # 复制模板工作表
                                template_sheet = workbook[template_sheet_name]
                                new_sheet = workbook.copy_worksheet(template_sheet)
                                new_sheet.title = sheet_title
                                
                                # 复制冻结窗格设置
                                if template_sheet.freeze_panes:
                                    new_sheet.freeze_panes = template_sheet.freeze_panes
                                
                                # 修改B列标题为【{msg_type}报文内容】
                                # 查找B列中包含'{$0}'的单元格作为标题
                                for row in range(1, min(11, new_sheet.max_row + 1)):
                                    cell_value = new_sheet[f'B{row}'].value
                                    if cell_value and '{$0}' in str(cell_value):
                                        original_title = cell_value
                                        new_title = original_title.replace('{$0}', msg_type)
                                        new_sheet[f'B{row}'].value = new_title
                                        break
                                found_sheet = new_sheet
                            else:
                                found_sheet = workbook.create_sheet(sheet_title)
                        
                        # 存储工作表对象，使用(类型,方向)作为键
                        message_sheets[(msg_type, direction)] = found_sheet
            
            # 为了向后兼容，保留原有的message_sheets[msg_type]引用
            for msg_type in supported_message_types:
                msg_name = supported_message_types[msg_type]
                # 默认使用unknown方向的工作表（如果存在）
                if (msg_type, 'unknown') in message_sheets:
                    message_sheets[msg_type] = message_sheets[(msg_type, 'unknown')]
                # 如果unknown方向不存在，检查recv或send方向是否存在
                elif (msg_type, 'recv') in message_sheets:
                    message_sheets[msg_type] = message_sheets[(msg_type, 'recv')]
                elif (msg_type, 'send') in message_sheets:
                    message_sheets[msg_type] = message_sheets[(msg_type, 'send')]
            
            # 按配置文件中的顺序排列工作表
            # 初始化prev_sheet_idx，从位置0开始（目录工作表会放在位置0）
            prev_sheet_idx = 0

            for msg_type in supported_message_types:
                msg_name = supported_message_types[msg_type]
                
                # 先处理recv方向的工作表（如果有）
                recv_sheet_name = f'{msg_name}_recv'
                if recv_sheet_name in workbook.sheetnames:
                    current_idx = workbook.sheetnames.index(recv_sheet_name)
                    if current_idx != prev_sheet_idx:
                        # 计算需要移动的偏移量
                        offset = prev_sheet_idx - current_idx
                        workbook.move_sheet(workbook[recv_sheet_name], offset=offset)
                    prev_sheet_idx += 1
                
                # 再处理send方向的工作表（如果有）
                send_sheet_name = f'{msg_name}_send'
                if send_sheet_name in workbook.sheetnames:
                    current_idx = workbook.sheetnames.index(send_sheet_name)
                    if current_idx != prev_sheet_idx:
                        # 计算需要移动的偏移量
                        offset = prev_sheet_idx - current_idx
                        workbook.move_sheet(workbook[send_sheet_name], offset=offset)
                    prev_sheet_idx += 1
                
                # 最后处理没有方向后缀的工作表（unknown方向）
                if msg_name in workbook.sheetnames:
                    current_idx = workbook.sheetnames.index(msg_name)
                    if current_idx != prev_sheet_idx:
                        # 计算需要移动的偏移量
                        offset = prev_sheet_idx - current_idx
                        workbook.move_sheet(workbook[msg_name], offset=offset)
                    prev_sheet_idx += 1
            
            # 创建目录sheet并放在最前面
            if '目录' in workbook.sheetnames:
                # 如果已存在目录sheet，先删除
                del workbook['目录']
            
            # 创建新的目录sheet
            catalog_sheet = workbook.create_sheet('目录', 0)  # 0表示放在最前面
            
            # 设置目录标题行
            catalog_sheet['A1'] = '序号'
            catalog_sheet['B1'] = '时间戳'
            catalog_sheet['C1'] = '查看详情'
            catalog_sheet['D1'] = '报文种类'
            catalog_sheet['E1'] = '传递方向'
            catalog_sheet['F1'] = '消息方向'
            
            # 设置标题行样式
            for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
                catalog_sheet[cell].font = openpyxl.styles.Font(bold=True)
            
            # 调整列宽
            catalog_sheet.column_dimensions['F'].width = 10
            
            # 创建时间戳到目录行号的映射，用于后续设置回目录链接
            timestamp_to_catalog_row = {}
            
            # 按时间戳排序所有报文
            all_messages.sort(key=lambda x: x['时间戳'])
            
            # 添加时间戳和链接到目录
            for row_idx, item in enumerate(all_messages, start=2):  # 从第2行开始
                # 记录时间戳到目录行号的映射
                timestamp_to_catalog_row[item['时间戳']] = row_idx
                
                # 在A列写入序号
                catalog_sheet[f'A{row_idx}'] = row_idx - 1
                # 在B列写入时间戳
                catalog_sheet[f'B{row_idx}'] = item['时间戳']
                # 添加链接指向对应的工作表中的行
                
                # 查找该报文在对应工作表中的行号
                msg_type = item['message_type']
                direction = item['方向']
                target_row = 3  # 默认第3行
                
                # 优先从messages_by_direction中按方向查找，确保找到正确方向的消息
                if msg_type in messages_by_direction and direction in messages_by_direction[msg_type]:
                    for i, msg in enumerate(messages_by_direction[msg_type][direction], start=3):
                        if msg['时间戳'] == item['时间戳']:
                            target_row = i
                            break
                # 如果按方向查找失败，再从processed_messages中查找
                elif msg_type in processed_messages:
                    for i, msg in enumerate(processed_messages[msg_type], start=3):
                        if msg['时间戳'] == item['时间戳'] and msg.get('方向') == direction:
                            target_row = i
                            break
                
                target_cell = f"'{item['sheet_name']}'!A{target_row}"
                catalog_sheet[f'C{row_idx}'] = "查看详情"
                catalog_sheet[f'C{row_idx}'].hyperlink = f"#{target_cell}"
                catalog_sheet[f'C{row_idx}'].font = openpyxl.styles.Font(color="0000FF", underline="single")
                
                # 添加报文种类
                catalog_sheet[f'D{row_idx}'] = item['报文种类']
                
                # 判断传递方向：55ee表示MCU->5G，ee55表示5G->MCU
                msg_type = item['message_type']
                content_key = f'{msg_type}报文内容'
                message_content = item.get(content_key, '')
                if message_content.startswith('55ee'):
                    direction = 'MCU->5G'
                elif message_content.startswith('ee55'):
                    direction = '5G->MCU'
                else:
                    direction = '未知方向'
                # 将传递方向写入目录sheet的E列
                catalog_sheet[f'E{row_idx}'] = direction
                
                # 添加消息方向(send/recv)到F列
                catalog_sheet[f'F{row_idx}'] = item.get('方向', '')
            
            # 定义行号替换函数一次，避免重复定义
            def replace_row_num(match, row_idx, source_row):
                # 检查是否是跨工作表引用
                if '!' in match.group(0):
                    return match.group(0)
                
                # 保留列引用部分
                col_part = match.group(1)
                # 获取原始行号
                original_row = int(match.group(2))
                # 判断是否是绝对行引用
                is_absolute_row = match.group(0).find(f'${original_row}') != -1
                
                # 如果不是绝对行引用，并且是引用的当前工作表中的行，则更新行号
                if not is_absolute_row:
                    # 计算行偏移量
                    row_diff = row_idx - source_row
                    # 如果原始行号是引用的模板中的数据行(>=3)，则更新行号
                    if original_row >= 3:
                        return f"{col_part}{original_row + row_diff}"
                # 对于绝对行引用或不是数据行的引用，保持不变
                return match.group(0)
            
            # 对所有报文类型和方向写入数据并应用公式
            for msg_type in messages_by_direction:
                for direction in messages_by_direction[msg_type]:
                    # 获取对应的工作表
                    if (msg_type, direction) not in message_sheets:
                        continue  # 如果没有对应的工作表，跳过
                    
                    sheet = message_sheets[(msg_type, direction)]
                    messages = messages_by_direction[msg_type][direction]
                    content_key = f'{msg_type}报文内容'
                    
                    if not messages:
                        continue  # 如果没有数据，跳过
                    
                    # 从A3和B3单元格开始写入数据
                    for row_idx, item in enumerate(messages, start=3):  # 从第3行开始
                        sheet[f'A{row_idx}'] = item['时间戳']   # A列写入时间戳
                        # 添加链接到目录sheet中的对应行
                        # 使用映射获取该时间戳在目录中的行号
                        if item['时间戳'] in timestamp_to_catalog_row:
                            catalog_row = timestamp_to_catalog_row[item['时间戳']]
                            sheet[f'A{row_idx}'].hyperlink = f"#目录!B{catalog_row}"
                        sheet[f'A{row_idx}'].font = openpyxl.styles.Font(color="0000FF", underline="single")
                        sheet[f'B{row_idx}'] = item[content_key]  # B列写入报文内容
                    
                    # 自动填充公式：复制C列及以后的公式到数据行
                    # 首先获取模板中已有的公式
                    # 动态获取模板的最大行数（从下往上查找最后一个有内容或公式的行）
                    max_template_row = sheet.max_row  # 获取工作表的最大行数
                    # 从max_row往上查找，找到最后一个有内容或公式的行
                    for row_idx in range(max_template_row, 2, -1):
                        has_content = False
                        for col_idx in range(1, min(20, sheet.max_column + 1)):  # 只检查前20列提高效率
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            # 检查单元格是否有值或包含公式
                            if cell.value is not None or (hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('=')):
                                has_content = True
                                break
                        if has_content:
                            max_template_row = row_idx
                            break
                    formulas = {}
                    
                    # 优化的公式模板获取逻辑
                    # 首先获取第3行的公式作为模板（数据起始行）
                    for col_idx in range(3, sheet.max_column + 1):
                        # 优先检查第3行（数据起始行）
                        cell = sheet.cell(row=3, column=col_idx)
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas[col_idx] = {
                                'formula': cell.value,
                                'source_row': 3,
                                'has_style': cell.has_style,
                                'font': copy.copy(cell.font) if cell.has_style else None,
                                'border': copy.copy(cell.border) if cell.has_style else None,
                                'fill': copy.copy(cell.fill) if cell.has_style else None,
                                'number_format': copy.copy(cell.number_format) if cell.has_style else None,
                                'protection': copy.copy(cell.protection) if cell.has_style else None,
                                'alignment': copy.copy(cell.alignment) if cell.has_style else None
                            }
                            continue
                        
                        # 如果第3行没有公式，从max_template_row行获取
                        if max_template_row >= 3:
                            cell = sheet.cell(row=max_template_row, column=col_idx)
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                formulas[col_idx] = {
                                    'formula': cell.value,
                                    'source_row': max_template_row,
                                    'has_style': cell.has_style,
                                    'font': copy.copy(cell.font) if cell.has_style else None,
                                    'border': copy.copy(cell.border) if cell.has_style else None,
                                    'fill': copy.copy(cell.fill) if cell.has_style else None,
                                    'number_format': copy.copy(cell.number_format) if cell.has_style else None,
                                    'protection': copy.copy(cell.protection) if cell.has_style else None,
                                    'alignment': copy.copy(cell.alignment) if cell.has_style else None
                                }
                    
                    # 对所有数据行应用公式，从第3行开始
                    # 优化后的公式应用逻辑，减少重复计算
                    if formulas:
                        # 对每个数据行应用公式
                        for row_idx in range(3, len(messages) + 3):
                            # 对每个有公式的列应用公式
                            for col_idx, formula_info in formulas.items():
                                formula = formula_info['formula']
                                source_row = formula_info['source_row']
                                
                                # 创建新单元格
                                new_cell = sheet.cell(row=row_idx, column=col_idx)
                                
                                # 创建lambda函数包装replace_row_num，传递row_idx和source_row
                                replace_func = lambda match: replace_row_num(match, row_idx, source_row)
                                
                                # 应用行号替换
                                new_formula = row_pattern.sub(replace_func, formula)
                                
                                # 设置新公式
                                new_cell.value = new_formula
                                
                                # 复制单元格样式（使用预缓存的样式）
                                if formula_info['has_style']:
                                    new_cell.font = formula_info['font']
                                    new_cell.border = formula_info['border']
                                    new_cell.fill = formula_info['fill']
                                    new_cell.number_format = formula_info['number_format']
                                    new_cell.protection = formula_info['protection']
                                    new_cell.alignment = formula_info['alignment']
            
            # 调整列宽
            catalog_sheet.column_dimensions['A'].width = 8
            catalog_sheet.column_dimensions['B'].width = 30
            catalog_sheet.column_dimensions['C'].width = 12
            catalog_sheet.column_dimensions['D'].width = 10
            catalog_sheet.column_dimensions['E'].width = 12
            
            # 删除未使用的工作表
            # 确定需要保留的工作表：目录 + 所有有实际数据的报文工作表 + 选项List
            used_sheet_names = {'目录', '选项List'}
            
            # 只收集那些有实际数据的工作表名称
            for msg_type in supported_message_types:
                for direction in ['recv', 'send', 'unknown']:
                    # 检查该方向是否有数据
                    if (msg_type in messages_by_direction and 
                        direction in messages_by_direction[msg_type] and 
                        messages_by_direction[msg_type][direction]):
                        # 获取对应的工作表名称
                        msg_name = supported_message_types[msg_type]
                        if direction == 'unknown':
                            sheet_name = msg_name
                        else:
                            sheet_name = f'{msg_name}_{direction}'
                        used_sheet_names.add(sheet_name)
            
            # 删除所有未使用的工作表，包括模板中可能存在的模板sheet（如{$0}报文）
            # 以及任何没有实际数据的工作表（如61报文_recv、2d报文_recv、67报文_send等）
            # 需要逆序删除，因为删除前面的sheet会改变后面sheet的索引
            sheets_to_delete = [name for name in workbook.sheetnames if name not in used_sheet_names]
            # 逆序删除工作表
            for sheet_name in reversed(sheets_to_delete):
                del workbook[sheet_name]
            
            # 保存为新的Excel文件
            workbook.save(excel_path)
        except Exception as e:
            print(f"使用模板时出错：{e}")
            print("将使用默认方式创建Excel文件")
            # 默认方式创建Excel文件
            import openpyxl
            workbook = openpyxl.Workbook()
            
            # 存储所有报文工作表
            message_sheets = {}
            
            # 遍历所有报文类型和方向，只创建有数据的工作表
            for msg_type in supported_message_types:
                msg_name = supported_message_types[msg_type]
                
                # 为每种报文类型的每个方向创建工作表，但只在该方向有数据时才创建
                for direction in ['recv', 'send', 'unknown']:
                    # 检查该方向是否有数据
                    if msg_type in messages_by_direction and direction in messages_by_direction[msg_type] and messages_by_direction[msg_type][direction]:
                        # 构建带方向后缀的工作表名称
                        if direction == 'unknown':
                            sheet_title = msg_name
                        else:
                            sheet_title = f'{msg_name}_{direction}'
                        
                        # 创建工作表
                        first_sheet_created = False
                        if sheet_title not in workbook.sheetnames:
                            if not workbook.sheetnames or workbook.sheetnames[0] == 'Sheet':
                                # 使用默认的Sheet作为第一个工作表
                                sheet = workbook[workbook.sheetnames[0]]
                                sheet.title = sheet_title
                                first_sheet_created = True
                            else:
                                sheet = workbook.create_sheet(sheet_title)
                        else:
                            sheet = workbook[sheet_title]
                        
                        # 存储工作表引用
                        message_sheets[(msg_type, direction)] = sheet
                        # 保持向后兼容，unknown方向的工作表作为默认引用
                        if direction == 'unknown':
                            message_sheets[msg_type] = sheet
            
            # 为了向后兼容，保留原有的message_sheets[msg_type]引用（如果前面没有设置）
            for msg_type in supported_message_types:
                # 如果还没有设置默认引用，检查是否有其他方向的工作表
                if msg_type not in message_sheets:
                    if (msg_type, 'unknown') in message_sheets:
                        message_sheets[msg_type] = message_sheets[(msg_type, 'unknown')]
                    elif (msg_type, 'recv') in message_sheets:
                        message_sheets[msg_type] = message_sheets[(msg_type, 'recv')]
                    elif (msg_type, 'send') in message_sheets:
                        message_sheets[msg_type] = message_sheets[(msg_type, 'send')]
            
            # 创建目录sheet并放在最前面
            if '目录' in workbook.sheetnames:
                # 如果已存在目录sheet，先删除
                del workbook['目录']
            
            # 创建新的目录sheet
            catalog_sheet = workbook.create_sheet('目录', 0)  # 0表示放在最前面
            
            # 设置目录标题行
            catalog_sheet['A1'] = '序号'
            catalog_sheet['B1'] = '时间戳'
            catalog_sheet['C1'] = '查看详情'
            catalog_sheet['D1'] = '报文种类'
            catalog_sheet['E1'] = '传递方向'
            
            # 设置标题行样式
            for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
                catalog_sheet[cell].font = openpyxl.styles.Font(bold=True)
            
            # 合并所有报文并按时间戳排序
            all_messages = []
            
            # 添加所有报文类型和方向到目录数据
            for msg_type in messages_by_direction:
                for direction in messages_by_direction[msg_type]:
                    msg_name = supported_message_types[msg_type]
                    # 构建带方向后缀的工作表名称
                    if direction == 'unknown':
                        sheet_title = msg_name
                    else:
                        sheet_title = f"{msg_name}_{direction}"
                    
                    content_key = f'{msg_type}报文内容'
                    messages = messages_by_direction[msg_type][direction]
                    
                    for item in messages:
                        all_messages.append({
                            '时间戳': item['时间戳'],
                            '报文内容': item.get(content_key, ''),
                            '报文种类': msg_type,
                            'sheet_name': sheet_title,
                            'message_type': msg_type,
                            'direction': direction
                        })
            
            # 按时间戳排序
            all_messages.sort(key=lambda x: x['时间戳'])
            
            # 创建时间戳到目录行号的映射
            timestamp_to_catalog_row = {}
            
            # 添加时间戳和链接到目录
            for row_idx, item in enumerate(all_messages, start=2):  # 从第2行开始
                # 记录时间戳到目录行号的映射
                timestamp_to_catalog_row[item['时间戳']] = row_idx
                
                # 在A列写入序号
                catalog_sheet[f'A{row_idx}'] = row_idx - 1
                # 在B列写入时间戳
                catalog_sheet[f'B{row_idx}'] = item['时间戳']
                # 添加链接指向对应的工作表中的行
                # 查找该报文在对应工作表中的行号
                msg_type = item['message_type']
                content_key = f'{msg_type}报文内容'
                target_row = None
                
                # 获取对应报文类型的处理后数据
                messages = processed_messages.get(msg_type, [])
                for i, msg in enumerate(messages, start=2):
                    if msg['时间戳'] == item['时间戳'] and content_key in msg:
                        target_row = i
                        break
                
                if target_row:
                    target_cell = f"'{item['sheet_name']}'!A{target_row}"
                    catalog_sheet[f'C{row_idx}'] = "查看详情"
                    catalog_sheet[f'C{row_idx}'].hyperlink = f"#{target_cell}"
                    catalog_sheet[f'C{row_idx}'].font = openpyxl.styles.Font(color="0000FF", underline="single")
                else:
                    catalog_sheet[f'C{row_idx}'] = "未找到"
                
                # 添加报文种类
                catalog_sheet[f'D{row_idx}'] = item['报文种类']
                
                # 判断传递方向：55ee表示MCU->5G，ee55表示5G->MCU
                message_content = item['报文内容']
                if message_content.startswith('55ee'):
                    direction = 'MCU->5G'
                elif message_content.startswith('ee55'):
                    direction = '5G->MCU'
                else:
                    direction = '未知方向'
                
                catalog_sheet[f'E{row_idx}'] = direction
            
            # 写入各类型报文数据
            for msg_type in messages_by_direction:
                msg_name = supported_message_types[msg_type]
                content_key = f'{msg_type}报文内容'
                
                for direction in messages_by_direction[msg_type]:
                    # 获取对应的工作表
                    if (msg_type, direction) not in message_sheets:
                        continue  # 如果没有对应的工作表，跳过
                    
                    sheet = message_sheets[(msg_type, direction)]
                    messages = messages_by_direction[msg_type][direction]
                    
                    if not messages:
                        continue  # 如果没有数据，跳过
                    
                    # 构建标题
                    if direction == 'unknown':
                        title = f'{msg_name}内容'
                    elif direction == 'recv':
                        title = f'{msg_name}_接收内容'
                    else:  # send
                        title = f'{msg_name}_发送内容'
                    
                    # 从第1行开始写入标题和数据
                    sheet['A1'] = '时间戳'
                    sheet['B1'] = title
                    sheet['A1'].font = openpyxl.styles.Font(bold=True)
                    sheet['B1'].font = openpyxl.styles.Font(bold=True)
                    
                    # 写入报文数据
                    for row_idx, item in enumerate(messages, start=2):
                        sheet[f'A{row_idx}'] = item['时间戳']
                        # 添加链接到目录sheet中的对应行
                        if item['时间戳'] in timestamp_to_catalog_row:
                            catalog_row = timestamp_to_catalog_row[item['时间戳']]
                            sheet[f'A{row_idx}'].hyperlink = f"#目录!B{catalog_row}"
                        sheet[f'A{row_idx}'].font = openpyxl.styles.Font(color="0000FF", underline="single")
                        sheet[f'B{row_idx}'] = item.get(content_key, '')
            
            # 调整列宽
            catalog_sheet.column_dimensions['A'].width = 8
            catalog_sheet.column_dimensions['B'].width = 30
            catalog_sheet.column_dimensions['C'].width = 12
            catalog_sheet.column_dimensions['D'].width = 10
            catalog_sheet.column_dimensions['E'].width = 12
            
            # 为每个报文工作表设置列宽
            for sheet in message_sheets.values():
                sheet.column_dimensions['A'].width = 30
                sheet.column_dimensions['B'].width = 100
            
            # 删除未使用的工作表
            # 确定需要保留的工作表：目录 + 所有有实际数据的报文工作表 + 选项List
            used_sheet_names = {'目录', '选项List'}
            
            # 只收集那些有实际数据的工作表名称
            for msg_type in supported_message_types:
                for direction in ['recv', 'send', 'unknown']:
                    # 检查该方向是否有数据
                    if (msg_type in messages_by_direction and 
                        direction in messages_by_direction[msg_type] and 
                        messages_by_direction[msg_type][direction]):
                        # 获取对应的工作表名称
                        msg_name = supported_message_types[msg_type]
                        if direction == 'unknown':
                            sheet_name = msg_name
                        else:
                            sheet_name = f'{msg_name}_{direction}'
                        used_sheet_names.add(sheet_name)
            
            # 删除所有未使用的工作表
            # 需要逆序删除，因为删除前面的sheet会改变后面sheet的索引
            sheets_to_delete = [name for name in workbook.sheetnames if name not in used_sheet_names]
            # 逆序删除工作表
            for sheet_name in reversed(sheets_to_delete):
                del workbook[sheet_name]
            
            # 保存Excel文件
            workbook.save(excel_path)
    


def run_parse_spi(workspace_dir, logs_dir, config_path, template_path):
    """
    执行 SPI 日志解析
    :param workspace_dir: 用户工作区根目录（Path 对象或 str）
    :param logs_dir: 相对于 workspace_dir 的日志文件夹路径，如 "parse_spi/logs"
    :param config_path: workspace_dir/spi_id.txt"
    :param template_path: workspace_dir/template.xlsx"
    :return: dict {"success": bool, "message": str, "output_path": str, "count": int, "types": list}
    """
    workspace_dir = Path(workspace_dir)
    # 确保用户 workspace 中有默认配置文件/模板，并获取实际路径
    user_config, user_template = ensure_user_has_default_files(workspace_dir)

    # 解析参数：如果传入的相对路径不存在，使用默认路径
    logs_abs = workspace_dir / logs_dir
    config_abs = workspace_dir / config_path if config_path else user_config
    if not config_abs.exists():
        config_abs = user_config
    template_abs = workspace_dir / template_path if template_path else user_template
    if not template_abs.exists():
        template_abs = user_template
    output_abs = workspace_dir

    # 检查必要的路径
    if not logs_abs.exists() or not logs_abs.is_dir():
        return {"success": False, "message": f"日志文件夹不存在: {logs_abs}"}
    if not config_abs.exists():
        return {"success": False, "message": f"配置文件不存在: {config_abs}"}
    if not template_abs.exists():
        return {"success": False, "message": f"模板文件不存在: {template_abs}"}

    # 动态加载报文类型配置
    supported_message_types, header_message_map = load_message_types(config_abs)
    if not supported_message_types:
        return {"success": False, "message": "配置文件无效或未找到任何报文类型"}

    # 收集所有 .log 文件
    log_files = list(logs_abs.glob("*.log"))
    if not log_files:
        return {"success": False, "message": f"日志文件夹中没有 .log 文件: {logs_abs}"}

    all_messages = {msg_type: [] for msg_type in supported_message_types}

    for log_file in log_files:
        for header, msg_types in header_message_map.items():
            for msg_type in msg_types:
                file_messages = extract_messages(str(log_file), msg_type, header)
                all_messages[msg_type].extend(file_messages)

    # 去重并排序
    result_dict = {}
    total_count = 0
    extracted_types = []
    for msg_type in supported_message_types:
        unique_messages = {}
        for msg in all_messages[msg_type]:
            if '时间' in msg:
                content_key = f'{msg_type}报文内容'
                unique_key = f"{msg['时间']}_{msg.get(content_key, '')}"
                unique_messages[unique_key] = msg
        result_list = list(unique_messages.values())
        result_list.sort(key=lambda x: datetime.strptime(x['时间'], '%Y-%m-%d %H:%M:%S.%f'))
        result_dict[msg_type] = result_list
        if result_list:
            total_count += len(result_list)
            extracted_types.append(msg_type)

    if total_count == 0:
        return {"success": False, "message": "未找到任何支持的报文数据"}

    # 生成输出文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if len(extracted_types) <= 3:
        excel_name = f"{'_'.join(extracted_types)}报文提取结果_{timestamp}.xlsx"
    else:
        excel_name = f"报文提取结果_{timestamp}.xlsx"
    output_abs.mkdir(parents=True, exist_ok=True)
    excel_path = output_abs / excel_name

    # 保存到 Excel
    save_to_excel(result_dict, str(excel_path), supported_message_types, str(template_abs))

    return {
        "success": True,
        "message": "解析完成",
        "output_path": str(excel_path),
        "count": total_count,
        "types": extracted_types
    }


if __name__ == "__main__":
    run_parse_spi()
